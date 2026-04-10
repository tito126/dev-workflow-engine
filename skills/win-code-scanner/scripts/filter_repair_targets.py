#!/usr/bin/env python3
"""
高危代码修复台账过滤工具

用途：
- 读取修复台账 Excel 工作表
- 自动识别表头行与关键列
- 跳过“是否需要修复”已明确标记为无需修复的记录
- 输出后续扫描应保留的目标记录与文件列表

示例：
    python scripts/filter_repair_targets.py "计划.xlsx" --sheet "病区护士（姚云）"
    python scripts/filter_repair_targets.py "计划.xlsx" --sheet "病区护士（姚云）" --files-out targets.txt --json-out targets.json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

HEADER_ALIASES = {
    "issue_desc": ["问题描述", "风险描述", "问题"],
    "file_path": ["涉及文件/范围", "涉及文件 / 范围", "涉及文件", "文件", "范围"],
    "line_no": ["行号", "代码行号", "所在行", "行数"],
    "status": ["是否需要修复", "是否需要整改", "是否需要改造", "是否修复", "是否处理"],
    "remark": ["备注", "处理备注", "备注说明"],
    "branch": ["git分支", "分支", "branch"],
}

SKIP_STATUS = {
    "不需要修复", "无需修复", "无须修复", "不修复",
    "不需要整改", "无需整改", "不整改",
    "不需要处理", "无需处理", "不处理",
    "否", "no", "n", "skip", "跳过"
}

ACTIVE_STATUS = {
    "需要修复", "需修复", "待修复", "待确认", "需要处理",
    "需要整改", "待整改", "待处理", "是", "yes", "y"
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text.replace("\u3000", " ")


def normalize_header(value: Any) -> str:
    text = normalize_text(value)
    for token in [" ", "\n", "\t", "（", "）", "(", ")"]:
        text = text.replace(token, "")
    return text.replace("／", "/")


def normalize_status(value: Any) -> str:
    text = normalize_text(value).lower()
    for token in [" ", "\n", "\t", "（", "）", "(", ")", "，", ",", "。", ".", "；", ";", "：", ":"]:
        text = text.replace(token, "")
    return text


def match_header(value: Any, aliases: Iterable[str]) -> bool:
    normalized = normalize_header(value)
    return any(normalized == normalize_header(alias) for alias in aliases)


def find_header_row(ws) -> Tuple[int, Dict[str, int]]:
    max_scan_rows = min(ws.max_row, 20)
    for row_idx in range(1, max_scan_rows + 1):
        mapping: Dict[str, int] = {}
        row_values = [ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1)]
        for field, aliases in HEADER_ALIASES.items():
            for col_idx, value in enumerate(row_values, start=1):
                if match_header(value, aliases):
                    mapping[field] = col_idx
                    break
        if "status" in mapping and "file_path" in mapping and ("issue_desc" in mapping or "line_no" in mapping):
            return row_idx, mapping
    raise ValueError("未在前 20 行内识别到修复台账表头，请检查工作表结构")


def classify_status(status_raw: str) -> str:
    normalized = normalize_status(status_raw)
    if normalized in SKIP_STATUS:
        return "skip"
    if normalized in ACTIVE_STATUS:
        return "active"
    if not normalized:
        return "blank"
    return "unknown"


def record_key(record: Dict[str, Any]) -> Tuple[str, str, str]:
    return (
        normalize_text(record.get("file_path")),
        normalize_text(record.get("line_no")),
        normalize_text(record.get("issue_desc")),
    )


def extract_records(xlsx_path: Path, sheet_name: Optional[str]) -> Dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表不存在: {sheet_name}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header_row, header_map = find_header_row(ws)

    all_records: List[Dict[str, Any]] = []
    selected_records: List[Dict[str, Any]] = []
    skipped_records: List[Dict[str, Any]] = []
    undecided_records: List[Dict[str, Any]] = []
    selected_keys = set()

    for row_idx in range(header_row + 1, ws.max_row + 1):
        issue_desc = normalize_text(ws.cell(row_idx, header_map.get("issue_desc", 0)).value if header_map.get("issue_desc") else "")
        file_path = normalize_text(ws.cell(row_idx, header_map["file_path"]).value)
        line_no = normalize_text(ws.cell(row_idx, header_map.get("line_no", 0)).value if header_map.get("line_no") else "")
        status_raw = normalize_text(ws.cell(row_idx, header_map["status"]).value)
        remark = normalize_text(ws.cell(row_idx, header_map.get("remark", 0)).value if header_map.get("remark") else "")
        branch = normalize_text(ws.cell(row_idx, header_map.get("branch", 0)).value if header_map.get("branch") else "")

        if not any([issue_desc, file_path, line_no, status_raw, remark, branch]):
            continue

        record = {
            "excel_row": row_idx,
            "issue_desc": issue_desc,
            "file_path": file_path,
            "line_no": line_no,
            "status": status_raw,
            "remark": remark,
            "branch": branch,
        }
        all_records.append(record)

        status_kind = classify_status(status_raw)
        record["status_kind"] = status_kind

        if status_kind == "skip":
            skipped_records.append(record)
            continue

        key = record_key(record)
        if key in selected_keys:
            continue
        selected_keys.add(key)
        selected_records.append(record)

        if status_kind in {"blank", "unknown"}:
            undecided_records.append(record)

    selected_files = sorted({r["file_path"] for r in selected_records if r.get("file_path")})
    skipped_files = sorted({r["file_path"] for r in skipped_records if r.get("file_path")})

    return {
        "xlsx_path": str(xlsx_path),
        "sheet_name": ws.title,
        "header_row": header_row,
        "header_map": header_map,
        "stats": {
            "total_records": len(all_records),
            "selected_records": len(selected_records),
            "skipped_no_fix_records": len(skipped_records),
            "undecided_records": len(undecided_records),
            "selected_files": len(selected_files),
            "skipped_files": len(skipped_files),
        },
        "selected_records": selected_records,
        "skipped_records": skipped_records,
        "undecided_records": undecided_records,
        "selected_files": selected_files,
        "skipped_files": skipped_files,
    }


def write_lines(path: Path, lines: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for line in lines:
            f.write(f"{line}\n")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="过滤修复台账中已标记无需修复的记录")
    parser.add_argument("xlsx_path", help="Excel 文件路径")
    parser.add_argument("--sheet", dest="sheet_name", help="工作表名称，不传则默认第一张")
    parser.add_argument("--json-out", help="将完整 JSON 结果写入文件")
    parser.add_argument("--files-out", help="将保留后的唯一文件列表写入文件")
    parser.add_argument("--selected-out", help="将保留后的记录 JSON 写入文件")
    parser.add_argument("--format", choices=["json", "text"], default="json", help="stdout 输出格式")
    return parser


def print_text_summary(result: Dict[str, Any]) -> None:
    stats = result["stats"]
    print("修复台账过滤结果")
    print("=" * 60)
    print(f"文件: {result['xlsx_path']}")
    print(f"工作表: {result['sheet_name']}")
    print(f"表头行: {result['header_row']}")
    print(f"总记录: {stats['total_records']}")
    print(f"保留记录: {stats['selected_records']}")
    print(f"跳过记录(无需修复): {stats['skipped_no_fix_records']}")
    print(f"待确认记录: {stats['undecided_records']}")
    print(f"保留文件数: {stats['selected_files']}")
    print(f"跳过文件数: {stats['skipped_files']}")
    print("=" * 60)

    preview = result["selected_records"][:10]
    if preview:
        print("保留记录预览:")
        for item in preview:
            line_suffix = f":{item['line_no']}" if item.get("line_no") else ""
            print(f"- 行{item['excel_row']}: {item['file_path']}{line_suffix} | {item['issue_desc']} | 状态={item['status'] or '空'}")


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path)
    if not xlsx_path.exists():
        print(f"错误: 文件不存在: {xlsx_path}", file=sys.stderr)
        return 1

    try:
        result = extract_records(xlsx_path, args.sheet_name)
    except Exception as exc:
        print(f"错误: {exc}", file=sys.stderr)
        return 1

    if args.json_out:
        output_path = Path(args.json_out)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")

    if args.files_out:
        write_lines(Path(args.files_out), result["selected_files"])

    if args.selected_out:
        output_path = Path(args.selected_out)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(json.dumps(result["selected_records"], ensure_ascii=False, indent=2), encoding="utf-8")

    if args.format == "text":
        print_text_summary(result)
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
