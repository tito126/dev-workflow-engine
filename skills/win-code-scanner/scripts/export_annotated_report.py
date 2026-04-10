#!/usr/bin/env python3
"""
将带知识库增强字段的扫描结果导出为 Markdown / Excel 报告。

支持输入：
- annotate_scan_results.py 的输出
- 或者已包含 kb_* 字段的 issues JSON

示例：
    python scripts/export_annotated_report.py annotated.json --xlsx-out report.xlsx --md-out report.md --repo-name winning-demo
"""

from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ACTION_DISPLAY = {
    "": "保留",
    "keep": "保留",
    "保留": "保留",
    "downgrade": "降级",
    "manual_review": "人工确认",
    "known_false_positive": "已知误报",
}

SEVERITY_DISPLAY = {
    "严重": "🔴 high",
    "警告": "🟡 medium",
    "提示": "🔵 low",
    "高": "🔴 high",
    "中": "🟡 medium",
    "低": "🔵 low",
}

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
TOP_ALIGNMENT = Alignment(vertical="top")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def load_issues(payload: Any) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    if isinstance(payload, dict):
        if isinstance(payload.get("issues"), list):
            return payload["issues"], payload
    if isinstance(payload, list):
        return payload, {"issues": payload}
    raise ValueError("输入 JSON 必须是 issues 数组，或包含 issues 数组的对象")


def issue_rule_code(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("rule_code") or issue.get("rule") or issue.get("ruleId"))


def issue_rule_name(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("rule_name") or issue.get("ruleName"))


def issue_file_path(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("file") or issue.get("file_path") or issue.get("path"))


def issue_line(issue: Dict[str, Any]) -> str:
    line = issue.get("line")
    if line is None:
        return normalize_text(issue.get("line_no"))
    return str(line)


def issue_description(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("description") or issue.get("issue_desc") or issue.get("message"))


def issue_suggestion(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("suggestion") or issue.get("fix_suggestion"))


def issue_severity(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("severity") or "警告")


def issue_category(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("category") or "未知")


def issue_code(issue: Dict[str, Any]) -> str:
    return normalize_text(issue.get("code_snippet") or issue.get("snippet") or issue.get("code"))


def kb_action_display(action: str) -> str:
    action = normalize_text(action)
    return ACTION_DISPLAY.get(action, action or "保留")


def severity_display(severity: str) -> str:
    severity = normalize_text(severity)
    return SEVERITY_DISPLAY.get(severity, severity or "🟡 medium")


def aggregate_rule_rows(issues: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    file_lines_by_rule: Dict[str, Dict[str, List[str]]] = defaultdict(lambda: defaultdict(list))

    for issue in issues:
        rule_code = issue_rule_code(issue) or "UNKNOWN"
        if rule_code not in grouped:
            grouped[rule_code] = {
                "rule_code": rule_code,
                "rule_name": issue_rule_name(issue),
                "severity": issue_severity(issue),
                "category": issue_category(issue),
                "suggestion": issue_suggestion(issue),
                "description": issue_description(issue),
                "kb_match_statuses": [],
                "kb_actions": [],
                "kb_rule_ids": [],
                "kb_reasons": [],
                "kb_feature_summaries": [],
                "kb_confidences": [],
                "kb_requires_human_review": [],
                "issues": [],
            }

        row = grouped[rule_code]
        row["issues"].append(issue)
        row["kb_match_statuses"].append(normalize_text(issue.get("kb_match_status")))
        row["kb_actions"].append(kb_action_display(issue.get("kb_action")))
        row["kb_rule_ids"].append(normalize_text(issue.get("kb_rule_id")))
        row["kb_reasons"].append(normalize_text(issue.get("kb_reason")))
        row["kb_feature_summaries"].append(normalize_text(issue.get("kb_feature_summary")))
        row["kb_confidences"].append(normalize_text(issue.get("kb_confidence")))
        row["kb_requires_human_review"].append(bool(issue.get("kb_requires_human_review", True)))

        file_path = issue_file_path(issue)
        line = issue_line(issue)
        if file_path:
            if line and line not in {"0", "None"}:
                file_lines_by_rule[rule_code][file_path].append(line)
            else:
                file_lines_by_rule[rule_code][file_path]

    output = []
    for rule_code, row in grouped.items():
        files_summary_lines = []
        impact_count = 0
        for file_path, lines in sorted(file_lines_by_rule[rule_code].items()):
            normalized_lines = []
            for line in lines:
                if line and line not in normalized_lines:
                    normalized_lines.append(line)
            impact_count += len(normalized_lines)
            if normalized_lines:
                files_summary_lines.append(f"{file_path}: 第 {', '.join(normalized_lines)} 行")
            else:
                files_summary_lines.append(file_path)

        def collapse(values: Iterable[Any], default: str = "") -> str:
            uniq = [normalize_text(v) for v in values if normalize_text(v)]
            uniq = list(dict.fromkeys(uniq))
            if not uniq:
                return default
            if len(uniq) == 1:
                return uniq[0]
            return " / ".join(uniq[:3]) if len(uniq) <= 3 else "混合"

        output.append({
            "rule_code": rule_code,
            "risk_level": severity_display(row["severity"]),
            "category": row["category"],
            "rule_name": row["rule_name"],
            "impact_count": impact_count,
            "files_summary": "\n".join(files_summary_lines),
            "suggestion": row["suggestion"],
            "description": row["description"],
            "kb_match_status": collapse(row["kb_match_statuses"], "未命中"),
            "kb_action": collapse(row["kb_actions"], "保留"),
            "kb_rule_id": collapse(row["kb_rule_ids"], ""),
            "kb_reason": collapse(row["kb_reasons"], ""),
            "kb_feature_summary": collapse(row["kb_feature_summaries"], ""),
            "kb_confidence": collapse(row["kb_confidences"], ""),
            "kb_requires_human_review": "是" if any(row["kb_requires_human_review"]) else "否",
        })

    severity_order = {"🔴 high": 0, "🟡 medium": 1, "🔵 low": 2}
    output.sort(key=lambda item: (severity_order.get(item["risk_level"], 9), item["rule_code"]))
    return output


def build_overview(repo_name: str, git_url: str, git_branch: str, issues: List[Dict[str, Any]]) -> Dict[str, Any]:
    severity_counter = Counter(issue_severity(issue) for issue in issues)
    action_counter = Counter(kb_action_display(issue.get("kb_action")) for issue in issues)
    match_counter = Counter(normalize_text(issue.get("kb_match_status")) or "未命中" for issue in issues)
    category_counter = Counter(issue_category(issue) for issue in issues)

    return {
        "repo_name": repo_name,
        "git_url": git_url,
        "git_branch": git_branch,
        "total_issues": len(issues),
        "critical": severity_counter.get("严重", 0) + severity_counter.get("高", 0),
        "warning": severity_counter.get("警告", 0) + severity_counter.get("中", 0),
        "notice": severity_counter.get("提示", 0) + severity_counter.get("低", 0),
        "action_counter": dict(action_counter),
        "match_counter": dict(match_counter),
        "category_counter": dict(category_counter),
    }


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 60))
            cell.alignment = WRAP_ALIGNMENT
        ws.column_dimensions[column_letter].width = max(12, min(max_length + 2, 60))


def style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGNMENT


def write_overview_sheet(wb: Workbook, overview: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "扫描概览"
    ws.append(["项目", "值"])
    style_header(ws[1])

    rows = [
        ["仓库", overview["repo_name"]],
        ["Git URL", overview["git_url"]],
        ["分支", overview["git_branch"]],
        ["问题总数", overview["total_issues"]],
        ["严重", overview["critical"]],
        ["警告", overview["warning"]],
        ["提示", overview["notice"]],
        ["知识库动作统计", "；".join(f"{k}:{v}" for k, v in overview["action_counter"].items())],
        ["知识库匹配统计", "；".join(f"{k}:{v}" for k, v in overview["match_counter"].items())],
        ["问题类别统计", "；".join(f"{k}:{v}" for k, v in overview["category_counter"].items())],
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    for cell in ws["A"]:
        cell.font = BOLD_FONT
        cell.alignment = TOP_ALIGNMENT
    for cell in ws["B"]:
        cell.alignment = WRAP_ALIGNMENT


def write_rule_sheet(wb: Workbook, rule_rows: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("问题清单")
    headers = [
        "规则编号", "风险等级", "问题类别", "规则名称", "影响范围", "涉及文件",
        "修复建议", "原理说明", "知识库匹配状态", "建议处理动作", "知识库规则ID",
        "知识库判断原因", "匹配特征摘要", "置信度", "是否需人工复核"
    ]
    ws.append(headers)
    style_header(ws[1])

    for row in rule_rows:
        ws.append([
            row["rule_code"], row["risk_level"], row["category"], row["rule_name"], row["impact_count"],
            row["files_summary"], row["suggestion"], row["description"], row["kb_match_status"],
            row["kb_action"], row["kb_rule_id"], row["kb_reason"], row["kb_feature_summary"],
            row["kb_confidence"], row["kb_requires_human_review"],
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def write_detail_sheet(wb: Workbook, issues: List[Dict[str, Any]]) -> None:
    ws = wb.create_sheet("问题明细")
    headers = [
        "规则编号", "规则名称", "风险等级", "问题类别", "文件路径", "行号", "命中代码片段",
        "原始问题描述", "原始修复建议", "知识库匹配状态", "建议处理动作", "知识库规则ID",
        "知识库判断原因", "匹配特征摘要", "置信度", "是否需人工复核"
    ]
    ws.append(headers)
    style_header(ws[1])

    for issue in issues:
        ws.append([
            issue_rule_code(issue),
            issue_rule_name(issue),
            severity_display(issue_severity(issue)),
            issue_category(issue),
            issue_file_path(issue),
            issue_line(issue),
            issue_code(issue),
            issue_description(issue),
            issue_suggestion(issue),
            normalize_text(issue.get("kb_match_status")) or "未命中",
            kb_action_display(issue.get("kb_action")),
            normalize_text(issue.get("kb_rule_id")),
            normalize_text(issue.get("kb_reason")),
            normalize_text(issue.get("kb_feature_summary")),
            normalize_text(issue.get("kb_confidence")),
            "是" if issue.get("kb_requires_human_review", True) else "否",
        ])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize_columns(ws)


def export_xlsx(path: Path, overview: Dict[str, Any], rule_rows: List[Dict[str, Any]], issues: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    write_overview_sheet(wb, overview)
    write_rule_sheet(wb, rule_rows)
    write_detail_sheet(wb, issues)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_markdown(path: Path, overview: Dict[str, Any], rule_rows: List[Dict[str, Any]], issues: List[Dict[str, Any]]) -> None:
    action_counter = overview["action_counter"]
    lines = [
        f"# {overview['repo_name']} 代码扫描报告（知识库增强版）",
        "",
        "## 扫描概览",
        "",
        f"- 仓库: `{overview['repo_name']}`",
        f"- Git URL: {overview['git_url']}",
        f"- 分支: `{overview['git_branch']}`",
        f"- 问题总数: `{overview['total_issues']}`",
        f"- 严重: `{overview['critical']}`，警告: `{overview['warning']}`，提示: `{overview['notice']}`",
        f"- 知识库动作统计: {'，'.join(f'{k} {v}' for k, v in action_counter.items()) if action_counter else '无'}",
        "",
        "## 按规则汇总",
        "",
        "| 规则编号 | 风险等级 | 问题类别 | 影响范围 | 建议处理动作 | 知识库规则ID | 是否需人工复核 |",
        "|---|---|---|---:|---|---|---|",
    ]
    for row in rule_rows:
        lines.append(
            f"| {row['rule_code']} | {row['risk_level']} | {row['category']} | {row['impact_count']} | {row['kb_action']} | {row['kb_rule_id'] or ''} | {row['kb_requires_human_review']} |"
        )

    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for issue in issues:
        groups[kb_action_display(issue.get("kb_action"))].append(issue)

    for action in ["保留", "人工确认", "降级", "已知误报"]:
        items = groups.get(action, [])
        if not items:
            continue
        lines.extend(["", f"## {action}", ""])
        for issue in items:
            lines.extend([
                f"### {issue_rule_code(issue)} - {issue_file_path(issue)}:{issue_line(issue)}",
                f"- 风险等级: {severity_display(issue_severity(issue))}",
                f"- 原始问题: {issue_description(issue)}",
                f"- 建议处理动作: {kb_action_display(issue.get('kb_action'))}",
                f"- 知识库匹配状态: {normalize_text(issue.get('kb_match_status')) or '未命中'}",
                f"- 知识库规则ID: {normalize_text(issue.get('kb_rule_id'))}",
                f"- 知识库判断原因: {normalize_text(issue.get('kb_reason'))}",
                f"- 匹配特征摘要: {normalize_text(issue.get('kb_feature_summary'))}",
                f"- 是否需人工复核: {'是' if issue.get('kb_requires_human_review', True) else '否'}",
                "",
            ])

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="导出知识库增强后的扫描报告")
    parser.add_argument("input_json", help="annotate_scan_results 的输出 JSON")
    parser.add_argument("--xlsx-out", help="导出 Excel 路径")
    parser.add_argument("--md-out", help="导出 Markdown 路径")
    parser.add_argument("--repo-name", default="code-scan", help="仓库名称")
    parser.add_argument("--git-url", default="", help="Git URL")
    parser.add_argument("--git-branch", default="", help="Git 分支")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input_json)
    if not input_path.exists():
        raise SystemExit(f"输入文件不存在: {input_path}")
    if not args.xlsx_out and not args.md_out:
        raise SystemExit("至少提供 --xlsx-out 或 --md-out 之一")

    payload = load_json(input_path)
    issues, _ = load_issues(payload)
    overview = build_overview(args.repo_name, args.git_url, args.git_branch, issues)
    rule_rows = aggregate_rule_rows(issues)

    if args.xlsx_out:
        export_xlsx(Path(args.xlsx_out), overview, rule_rows, issues)
    if args.md_out:
        export_markdown(Path(args.md_out), overview, rule_rows, issues)

    print(json.dumps({
        "total_issues": len(issues),
        "xlsx_out": args.xlsx_out or "",
        "md_out": args.md_out or "",
        "rule_rows": len(rule_rows),
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
